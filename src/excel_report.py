"""
excel_report.py

Exportação e formatação corporativa do "data warehouse" multi-abas.
Antes este código vivia inline na última célula do notebook 06; foi extraído
para um módulo puro e reutilizável (notebooks/scripts apenas orquestram).
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta corporativa ────────────────────────────────────────────────────────
VERDE = "00925C"
AZUL = "08484C"
VERDE_LIMA = "B0CD4E"
LARANJA = "EF7925"
BRANCO = "FFFFFF"
CINZA_ALT = "F4F4F4"
CINZA_BORDA = "DCDCDC"

# Cor do cabeçalho por aba (alterna entre as 4 cores da paleta)
SHEET_HEADER_COLORS = {
    "1_Visao_Clientes_CRM": VERDE,
    "2_Demografia_E_Taxas": AZUL,
    "3_CurvaABC_Rentabilidade": VERDE,
    "4_Matriz_CrossSell": LARANJA,
    "5_Cohort_Sazonalidade": AZUL,
    "6_Market_Share_Fornecedor": VERDE_LIMA,
    "7_Calculadora_Produtos": VERDE,
    "8_AUDITORIA_BRUTA_LASTRO": LARANJA,
    "9_LOG_QUALIDADE_BACKOFFICE": AZUL,
    "10_Mix_Por_Especialidade": AZUL,
    "11_Agenda_Renovacoes_90d": LARANJA,
    "12_Win_Back_Reativacao": VERDE,
    "13_Snapshot_Mensal_Ativo": AZUL,
    "14_CurvaABC_2025plus": VERDE,
    "15_Market_Share_2025plus": VERDE_LIMA,
    "16_CurvaABC_Comissao": LARANJA,
    "17_Margem_Comissao_Seguradora": LARANJA,
    "18_Margem_Comissao_Produto": LARANJA,
    # Apoio não-monetário (abas distribuídas nos workbooks Comercial/Operacional)
    "0_Completude_Cadastro": AZUL,
    "1_Mix_Produtos": VERDE,
    "2_Market_Share_Contagem": VERDE_LIMA,
    "3_Profundidade_Carteira": VERDE,
    "4_Taxa_Cancelamento": LARANJA,
    "5_CrossSell_Gaps_Especialidade": LARANJA,
    # Arquivo de qualidade DQ (DQ_Reports.xlsx)
    "0_RESUMO_DQ": AZUL,
    "1_PREMIO_ZERADO_NEGATIVO": LARANJA,
    "2_COMISSAO_MAIOR_PREMIO": LARANJA,
    "3_INCONSISTENCIA_PERCENTUAL": LARANJA,
    "4_OUTLIER_PREMIO": VERDE,
    "5_OUTLIER_PCT_COMISSAO": VERDE,
    "6_GRUPOS_AMOSTRA_INSUF": VERDE_LIMA,
    "7_DUPLICATAS_EXATAS": LARANJA,
    "8_LOG_QUALIDADE_GERAL": AZUL,
    "9_AUDITORIA_BRUTA_LASTRO": AZUL,
}


def get_header_font_color(bg_color):
    """Retorna preto para fundos claros (verde-lima), branco para escuros."""
    return "1A1A1A" if bg_color == VERDE_LIMA else BRANCO


def format_sheet(ws, header_color):
    """
    Aplica formatação corporativa em uma aba:
    - Cabeçalho com cor da paleta, negrito, fonte Arial
    - Largura auto-ajustada por conteúdo (mín 10, máx 45)
    - Listras zebra em cinza claro nas linhas de dados
    - Bordas discretas em toda a tabela
    - Linha 1 congelada + filtro automático
    """
    font_color = get_header_font_color(header_color)

    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(name="Arial", bold=True, color=font_color, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill = PatternFill("solid", fgColor=CINZA_ALT)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color=CINZA_BORDA)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column
    max_row = ws.max_row

    # Cabeçalho
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Linhas de dados
    for row in range(2, max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    # Auto-ajuste de largura
    for col in range(1, max_col + 1):
        max_length = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(max_length + 2, 10), 45
        )

    # Altura fixa do cabeçalho (acomoda wrap_text)
    ws.row_dimensions[1].height = 32

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_formatted_workbook(sheet_exports, output_file, header_colors=None):
    """
    Escreve uma lista de (nome_aba, DataFrame) em um único Excel e aplica a
    formatação corporativa em cada aba.

    sheet_exports : list[tuple[str, pd.DataFrame]]
    output_file   : caminho do .xlsx de saída
    header_colors : dict aba->cor; ausências usam VERDE como padrão.
    """
    if header_colors is None:
        header_colors = SHEET_HEADER_COLORS

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, df in sheet_exports:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(output_file)
        for sheet_name, _ in sheet_exports:
            if sheet_name in wb.sheetnames:
                format_sheet(wb[sheet_name], header_colors.get(sheet_name, VERDE))
        wb.save(output_file)
    except PermissionError:
        # Arquivo aberto no Excel (lock do Windows): avisa e segue sem derrubar o run
        print(
            f"  [BLOQUEADO] {Path(output_file).name} está aberto/protegido — "
            f"pulei a gravação. Feche o arquivo e rode novamente."
        )
        return None

    return output_file
